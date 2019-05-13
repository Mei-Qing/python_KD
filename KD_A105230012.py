#!/usr/bin/env python
# coding: utf-8

# In[24]:


import openpyxl
import pandas as pd
import os
import datetime
#from IPython.display import display


# In[25]:


def print_full(x):
    df = pd.DataFrame(x)
    pd.set_option('display.max_rows', 2000)
    pd.set_option('display.max_columns', 20)
    display(df)


# In[26]:


excelname= "kd"
sheetname = "kd"

wb = openpyxl.load_workbook(str(excelname)+".xlsx")
ws = wb[str(sheetname)]

dataset = []

i = 1
while(ws.cell(row=i, column=1).value!=None):
    row = []
    j = 1
    while(ws.cell(row=i, column=j).value!=None):
        row.append(ws.cell(row=i, column=j).value)
        j = j +1
    dataset.append(row)
    i = i + 1


# In[27]:


pd.DataFrame(dataset)


# In[28]:


trade=[]
for i in range(0,len(dataset)):
    row=[]
    for j in range(0,11):
        row.append(0)
    trade.append(row)
trade[0][0] = "date"
trade[0][1] = "open"
trade[0][2] = "high"
trade[0][3] = "low"
trade[0][4] = "close"
trade[0][5] = "rsv"
trade[0][6] = "k"
trade[0][7] = "d"
trade[0][8] = "buy_day"
trade[0][9] = "cum_day"
trade[0][10] = "sell_day"


# In[29]:


#抓KD數據
for i in range(1, len(trade)):
    trade[i][0] = dataset[i][0]
    trade[i][1] = float(dataset[i][1])
    trade[i][2] = float(dataset[i][2])
    trade[i][3] = float(dataset[i][3])
    trade[i][4] = float(dataset[i][4])



#rsv
#RSV(t) = (close(t) - min(low(z)| z = t, t-1, ..., t-8)/(max(high(z)) - min(low(z)| z = t, t-1, ..., t-8)*100 , t = 9, 11, ..., T
for i in range(9, len(trade)):
    rsv = 0
    high = []
    low = []
    for j in range(0, 9):
        high.append(trade[i-j][2])
        low.append(trade[i-j][3])
    rsv = (trade[i][4]-min(low))/(max(high)-min(low))*100
    trade[i][5] = rsv
    print(rsv)


# In[30]:


#k (K值 初始值 = 50, at day 9，所以 K(9) = 50)
trade[9][6] = 50

for i in range(10,len(trade)):
    K = 0 
    K = (trade[i-1][6]*2/3) + (trade[i][5]*1/3)
    trade[i][6] = K
    print(K)


# In[31]:


#d (D值 初始值 = 50, at day 9，所以 D(9) = 50)
trade[9][7] = 50

for i in range(10,len(trade)):
    D = 0 
    D =(trade[i-1][7]*2/3) + (trade[i][6]*1/3) #trade[i][6]不能直接寫成K，因為寫K會只抓到最後一筆K值
    trade[i][7] = D
    print(D)


# In[32]:


#KD黃金交叉
#當KD指標的K值由下往上突破D值，建議買進、做多
#K(t) > D(t)
for i in range(10, len(trade)):
    if(trade[i][6] > trade[i][7])and(trade[i-1][9]==0):
        trade[i][8] = 1
        trade[i][9] = 1
        trade[i][10] = 0
    elif(trade[i][6] < trade[i][7])and(trade[i-1][9]==1):
        trade[i][8] = 0
        trade[i][9] = 0
        trade[i][10] = 1
    elif(trade[i-1][2]>0):
        trade[i][8] = 0
        trade[i][9] = 1
        trade[i][10] = 0


# In[33]:


for i in range(1, len(trade)):
    trade[i][0] = dataset[i][0]     #把日期、股價給trade
    trade[i][7] = float(dataset[i][7])


# In[34]:


for i in range(0, len(trade)):
    for j in range(0, len(trade[1])):
        ws.cell(row=i+1, column=j+1).value= trade[i][j]
wb.save('kd3.xlsx')


# In[ ]:





# In[ ]:




